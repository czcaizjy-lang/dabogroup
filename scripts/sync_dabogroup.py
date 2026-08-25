#!/usr/bin/env python3
"""
抖音达播看板数据提取脚本（独立项目）
直接读取上游 Excel，生成 data/dabogroup_data.json

数据口径：
  - 全量达人 = 5 个商务 sheet 名单 ∪ 日流水全部达人
  - 归属商务：按优先级 久酒>雅宁>奥易>檀雅>星辞 匹配，无归属归「其他」
    （星辞业绩 sheet 即全量总花名册，故排最后，只保留专属它的达人）
  - 月度指标按当月聚合；趋势图取近 30 天滚动窗口（单位：万）
"""

import json
import os
import subprocess as sp
from collections import defaultdict
from datetime import datetime, timedelta
from openpyxl import load_workbook

# 直接读取上游文件
XLSX_PATH = '/Users/xiaocao/Desktop/蕉下文件/业绩追击/by月业绩/6月业绩/6月业绩追击（纯直播）.xlsx'
OUTPUT_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'dabogroup_data.json')

ZB_PRECEDENCE = ['久酒', '雅宁', '奥易', '檀雅', '星辞']   # 重叠归属优先级（星辞=全量花名册，排最后）
ZB_ORDER = ['星辞', '雅宁', '久酒', '奥易', '檀雅', '其他']
TREND_DAYS = 30  # 趋势图近 N 天


def find_live_sheet(wb):
    """自动查找日流水 sheet（名称包含「直播数据」）"""
    for name in wb.sheetnames:
        if '直播数据' in name:
            return name
    raise ValueError(f'未找到直播数据 sheet，可用 sheets: {wb.sheetnames}')


def safe_float(v):
    try:
        return float(v) if v is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def run():
    wb = load_workbook(XLSX_PATH, data_only=True)

    # ═══════════════════════════════════════════
    # === 1. 商务名单（久酒 / 雅宁 / 星辞 / 奥易 / 檀雅）===
    # ═══════════════════════════════════════════
    person_sheet_names = ['久酒业绩', '雅宁业绩', '星辞业绩', '奥易业绩', '檀雅业绩']
    person_douyin_sets = {}   # person_name -> set(douyin_ids)
    person_name_map = {}      # douyin_id -> 昵称（从人员 sheet A 列）

    for sheet_name in person_sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        douyin_ids = set()
        for r in range(2, ws.max_row + 1):
            did = ws.cell(r, 2).value  # B列
            if did:
                did_str = str(did)
                douyin_ids.add(did_str)
                name = ws.cell(r, 1).value  # A列
                if name:
                    person_name_map[did_str] = str(name)
        if douyin_ids:
            person_name = sheet_name.replace('业绩', '')
            person_douyin_sets[person_name] = douyin_ids

    # 星辞业绩花名册（昵称兜底来源）
    all_roster = {}  # douyin_id → {name}
    if '星辞业绩' in wb.sheetnames:
        ws_anchor = wb['星辞业绩']
        for r in range(3, ws_anchor.max_row + 1):
            douyin_id = ws_anchor.cell(r, 2).value
            if not douyin_id:
                continue
            all_roster[str(douyin_id)] = {
                '主播昵称': str(ws_anchor.cell(r, 1).value or ''),
            }

    # ═══════════════════════════════════════════
    # === 2. 读取日流水（X月直播数据）===
    # ═══════════════════════════════════════════
    live_sheet_name = find_live_sheet(wb)
    ws_live = wb[live_sheet_name]
    print(f'  日流水 sheet: {live_sheet_name}')

    daily_gmv = defaultdict(lambda: defaultdict(float))        # douyin_id → date → gmv
    daily_paid = defaultdict(lambda: defaultdict(float))      # douyin_id → date → paid
    daily_refund = defaultdict(lambda: defaultdict(float))    # douyin_id → date → refund
    daily_ad = defaultdict(lambda: defaultdict(float))        # douyin_id → date → ad_cost
    daily_commission = defaultdict(lambda: defaultdict(float)) # douyin_id → date → commission
    daily_duration = defaultdict(lambda: defaultdict(float))   # douyin_id → date → duration(min)
    live_name_map = {}  # 抖音号 → 昵称（日流水覆盖最全）

    all_dates_set = set()

    for r in range(2, ws_live.max_row + 1):
        douyin_id_raw = ws_live.cell(r, 3).value  # C列
        nickname_raw = ws_live.cell(r, 2).value    # B列
        dt_val = ws_live.cell(r, 4).value           # D列

        if not douyin_id_raw or not dt_val:
            continue

        douyin_id = str(douyin_id_raw)
        date_key = str(dt_val)[:10].replace('/', '-')
        all_dates_set.add(date_key)

        if nickname_raw:
            live_name_map[douyin_id] = str(nickname_raw)

        gmv = safe_float(ws_live.cell(r, 26).value)           # Z: 直播间成交金额
        paid = safe_float(ws_live.cell(r, 27).value)           # AA: 支付金额
        refund = safe_float(ws_live.cell(r, 32).value)         # AF: 退款金额
        ad_bind = safe_float(ws_live.cell(r, 44).value)        # AR: 投放消耗(绑定)
        ad_beitou = safe_float(ws_live.cell(r, 45).value)      # AS: 投放消耗(被投)
        commission = safe_float(ws_live.cell(r, 34).value)     # AH: 预估佣金支出
        duration = safe_float(ws_live.cell(r, 6).value)        # F: 直播时长(分钟)

        # 消耗：被投优先，0 回退绑定（与原看板一致）
        ad_cost = ad_beitou if ad_beitou > 0 else ad_bind

        daily_gmv[douyin_id][date_key] += gmv
        daily_paid[douyin_id][date_key] += paid
        daily_refund[douyin_id][date_key] += refund
        daily_ad[douyin_id][date_key] += ad_cost if gmv > 0 else 0
        daily_commission[douyin_id][date_key] += commission
        daily_duration[douyin_id][date_key] += duration

    all_dates = sorted(all_dates_set)

    # ═══════════════════════════════════════════
    # === 3. 确定当前月份 ===
    # ═══════════════════════════════════════════
    latest_date = all_dates[-1]
    current_month = latest_date[:7]  # e.g. "2026-07"
    current_month_dates = [d for d in all_dates if d.startswith(current_month)]
    print(f'  当前月份: {current_month} | 最新日期: {latest_date} | 当月天数: {len(current_month_dates)}')

    def sum_month(daily_dict, douyin_id):
        """汇总某达人当月所有日期的值"""
        return round(sum(
            daily_dict.get(douyin_id, {}).get(d, 0)
            for d in current_month_dates
        ), 2)

    def count_active_days(daily_gmv_dict, douyin_id):
        """统计当月 GMV > 0 的天数"""
        return sum(1 for d in current_month_dates
                   if daily_gmv_dict.get(douyin_id, {}).get(d, 0) > 0)

    # ═══════════════════════════════════════════
    # === 4. 归属商务映射（全量达人）===
    # ═══════════════════════════════════════════
    zb_of = {}              # douyin_id -> 归属商务
    zb_ids = {zb: set() for zb in ZB_ORDER}   # 商务名 -> set(douyin_ids)

    union_ids = set()
    for _pn, _ids in person_douyin_sets.items():
        union_ids |= _ids
    union_ids |= set(daily_gmv.keys())
    union_ids |= set(daily_paid.keys())

    for douyin_id in union_ids:
        matched = None
        for pn in ZB_PRECEDENCE:
            if douyin_id in person_douyin_sets.get(pn, set()):
                matched = pn
                break
        if matched is None:
            matched = '其他'
        zb_of[douyin_id] = matched
        zb_ids[matched].add(douyin_id)

    print('  归属商务: ' + ' | '.join(
        f'{zb}={len(zb_ids[zb])}人' for zb in ZB_ORDER) + f' | 全量={len(union_ids)}人')

    # ═══════════════════════════════════════════
    # === 5. 全量达人月度指标（含归属商务）===
    # ═══════════════════════════════════════════
    zb_anchor_monthly = {}
    for douyin_id in union_ids:
        gmv_val = sum_month(daily_gmv, douyin_id)
        paid_val = sum_month(daily_paid, douyin_id)
        refund_val = sum_month(daily_refund, douyin_id)
        ad_val = sum_month(daily_ad, douyin_id)
        commission_val = sum_month(daily_commission, douyin_id)
        duration_total = sum_month(daily_duration, douyin_id)
        active_days = count_active_days(daily_gmv, douyin_id)
        settle_val = round(paid_val - refund_val, 2)
        info = all_roster.get(douyin_id, {})
        name = live_name_map.get(douyin_id) or person_name_map.get(douyin_id) or info.get('主播昵称') or douyin_id
        zb_anchor_monthly[douyin_id] = {
            '主播昵称': str(name),
            '主播抖音号': douyin_id,
            '归属商务': zb_of[douyin_id],
            '直播GMV': gmv_val,
            '直播支付GMV': paid_val,
            '直播退款GMV': refund_val,
            '直播结算GMV': settle_val,
            '结算率': round(settle_val / gmv_val, 4) if gmv_val > 0 else 0,
            'ROI': round(gmv_val / ad_val, 2) if ad_val > 0 else 0,
            '佣金支出': commission_val,
            '投放消耗金额': ad_val,
        }

    # ═══════════════════════════════════════════
    # === 6. 各商务汇总（含全部）===
    # ═══════════════════════════════════════════
    def _zb_summary(dids):
        gmv = sum(zb_anchor_monthly[d]['直播GMV'] for d in dids)
        paid = sum(zb_anchor_monthly[d]['直播支付GMV'] for d in dids)
        refund = sum(zb_anchor_monthly[d]['直播退款GMV'] for d in dids)
        settle = round(paid - refund, 2)
        ad = sum(zb_anchor_monthly[d]['投放消耗金额'] for d in dids)
        commission = sum(zb_anchor_monthly[d]['佣金支出'] for d in dids)
        return {
            '直播GMV': round(gmv, 2),
            '直播支付GMV': round(paid, 2),
            '直播退款GMV': round(refund, 2),
            '直播结算GMV': settle,
            '结算率': round(settle / gmv, 4) if gmv > 0 else 0,
            '佣金支出': round(commission, 2),
            '投放消耗金额': round(ad, 2),
            '达人数': len(dids),
        }

    summary_by_zb = {}
    for zb in ZB_ORDER:
        summary_by_zb[zb] = _zb_summary(list(zb_ids[zb]))
    summary_by_zb['全部'] = _zb_summary(list(union_ids))

    # ═══════════════════════════════════════════
    # === 7. 趋势数据（近 30 天）===
    # ═══════════════════════════════════════════
    trend_dates_full = all_dates[-TREND_DAYS:] if len(all_dates) >= TREND_DAYS else all_dates
    trend_dates = [f"{int(d[5:7])}/{int(d[8:10])}" for d in trend_dates_full]
    date_map = {trend_dates_full[i]: trend_dates[i] for i in range(len(trend_dates_full))}

    # 全量每日聚合（所有日流水达人）
    total_gmv_daily = defaultdict(float)
    total_paid_daily = defaultdict(float)
    total_refund_daily = defaultdict(float)

    for d in trend_dates_full:
        for douyin_id, by_date in daily_gmv.items():
            total_gmv_daily[d] += by_date.get(d, 0)
            total_paid_daily[d] += daily_paid[douyin_id].get(d, 0)
            total_refund_daily[d] += daily_refund[douyin_id].get(d, 0)

    def daily_list(daily_dict):
        """将 defaultdict 转为按 trend_dates_full 顺序的列表（单位：万）"""
        return [round(daily_dict.get(d, 0) / 10000, 2) for d in trend_dates_full]

    # 各商务每日趋势（万）
    def _zb_daily_list(daily_dict, dids):
        return [round(sum(daily_dict.get(did, {}).get(d, 0) for did in dids) / 10000, 2) for d in trend_dates_full]

    daily_by_zb = {}
    daily_paid_by_zb = {}
    daily_refund_by_zb = {}
    for zb in ZB_ORDER:
        dids = list(zb_ids[zb])
        daily_by_zb[zb] = _zb_daily_list(daily_gmv, dids)
        daily_paid_by_zb[zb] = _zb_daily_list(daily_paid, dids)
        daily_refund_by_zb[zb] = _zb_daily_list(daily_refund, dids)

    # 分达人下探数据
    anchor_daily_paid_out = {}
    for douyin_id in daily_paid:
        anchor_daily_paid_out[douyin_id] = {
            date_map[d]: round(daily_paid[douyin_id].get(d, 0) / 10000, 2)
            for d in trend_dates_full
        }

    anchor_daily_roi = {}
    for douyin_id in daily_gmv:
        anchor_daily_roi[douyin_id] = {}
        for d in trend_dates_full:
            g = daily_gmv[douyin_id].get(d, 0)
            a = daily_ad.get(douyin_id, {}).get(d, 0)
            anchor_daily_roi[douyin_id][date_map[d]] = round(g / a, 2) if a > 0 else 0

    # ═══════════════════════════════════════════
    # === 8. 达人列表 / 下探 / Top5 ===
    # ═══════════════════════════════════════════
    zb_anchors_sorted = sorted(zb_anchor_monthly.values(), key=lambda x: x['直播GMV'], reverse=True)

    all_anchor_daily_zb = []
    for douyin_id in daily_gmv:
        info = all_roster.get(douyin_id, {})
        name = live_name_map.get(douyin_id) or person_name_map.get(douyin_id) or info.get('主播昵称') or douyin_id
        vals = [round(daily_paid.get(douyin_id, {}).get(d, 0) / 10000, 2) for d in trend_dates_full]
        all_anchor_daily_zb.append({
            'name': str(name),
            'douyin_id': douyin_id,
            'zb': zb_of.get(douyin_id, '其他'),
            'daily_paid': vals,
        })
    all_anchor_daily_zb.sort(key=lambda x: sum(x['daily_paid']), reverse=True)

    zb_top5 = {}
    for zb in ZB_ORDER:
        aids = sorted(
            [d for d in zb_ids[zb] if zb_anchor_monthly[d]['直播GMV'] > 0],
            key=lambda d: zb_anchor_monthly[d]['直播GMV'], reverse=True
        )[:5]
        top5 = []
        for douyin_id in aids:
            info = all_roster.get(douyin_id, {})
            name = live_name_map.get(douyin_id) or person_name_map.get(douyin_id) or info.get('主播昵称') or douyin_id
            vals = [round(daily_paid.get(douyin_id, {}).get(d, 0) / 10000, 2) for d in trend_dates_full]
            top5.append({'name': str(name), 'douyin_id': douyin_id, 'daily_paid': vals})
        zb_top5[zb] = top5

    # ═══════════════════════════════════════════
    # === 9. 构建最终 JSON ===
    # ═══════════════════════════════════════════
    dabogroup_data = {
        'trend': {
            'dates': trend_dates,
            'daily_total_gmv': daily_list(total_gmv_daily),
            'daily_total_paid': daily_list(total_paid_daily),
            'daily_total_refund': daily_list(total_refund_daily),
            'anchor_daily_paid': anchor_daily_paid_out,
            'anchor_daily_roi': anchor_daily_roi,
        },
        'dabogroup': {
            'anchors': zb_anchors_sorted,
            'summary_by_zb': summary_by_zb,
            'daily_by_zb': daily_by_zb,
            'daily_paid_by_zb': daily_paid_by_zb,
            'daily_refund_by_zb': daily_refund_by_zb,
            'all_anchor_daily': all_anchor_daily_zb,
            'top5_by_zb': zb_top5,
        },
    }

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(dabogroup_data, f, ensure_ascii=False, indent=2)

    print(f'✓ 数据已更新到 {OUTPUT_PATH}')
    print(f'  - 当前月份: {current_month}')
    print(f'  - 趋势窗口: {trend_dates[0]} ~ {trend_dates[-1]} ({len(trend_dates)} 天)')
    print(f'  - 全量达人: {len(zb_anchors_sorted)}')
    print(f'  - 全部 GMV: ¥{summary_by_zb["全部"]["直播GMV"]:,.2f}')
    print('  - 商务分布: ' + ' | '.join(
        f'{zb}={summary_by_zb[zb]["达人数"]}人' for zb in ZB_ORDER))

    # 自动构建独立看板页面
    script_dir = os.path.dirname(os.path.abspath(__file__))
    build_script = os.path.join(script_dir, 'build_dabogroup_standalone.py')
    if os.path.exists(build_script):
        print('')
        sp.run(['python3', build_script], cwd=script_dir)


if __name__ == '__main__':
    run()
